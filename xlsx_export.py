import io
import json
import math
from typing import Any, Dict, List, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
RECOMMENDED_FILL = PatternFill(fill_type="solid", fgColor="F3F6FA")
EDITABLE_FILL = PatternFill(fill_type="solid", fgColor="FFF4CC")

DEFAULT_OVERRIDE_FIELDS: List[Dict[str, str]] = [
    {
        "field": "global_renaming_instructions",
        "description": "Describe global renaming decisions, including friendlier names for identifiers, dates, question blocks, or child tables.",
    },
    {
        "field": "global_regex_rules",
        "description": "Provide any regex-style grouping or naming rules that should be applied consistently across the dataset.",
    },
    {
        "field": "missed_family_information",
        "description": "List repeat or matrix families that were missed, plus any known business labels for their row indices.",
    },
    {
        "field": "free_text_override_instructions",
        "description": "Capture any downstream structural instructions, such as fields to keep wide, tables to build, or columns that should never be dropped.",
    },
]


def _normalize_rows(obj: Any) -> Tuple[List[str], List[List[Any]]]:
    """
    Accepts:
      - list[dict]  -> headers from keys (deterministic order)
      - dict with {"columns": [...], "rows": [...]}
      - list[list]  -> assumes first row is header if all strings, else generates Col1..ColN
    Returns (headers, rows)
    """
    if isinstance(obj, dict) and "columns" in obj and "rows" in obj:
        cols = [str(c) for c in obj["columns"]]
        rows = obj["rows"]
        if not isinstance(rows, list):
            raise ValueError("rows must be a list")
        return cols, [list(r) if isinstance(r, (list, tuple)) else [r] for r in rows]

    if isinstance(obj, list) and obj and isinstance(obj[0], dict):
        first_keys = list(obj[0].keys())
        extra_keys = sorted({k for r in obj[1:] for k in (r.keys() if isinstance(r, dict) else [])} - set(first_keys))
        headers = [str(k) for k in (first_keys + extra_keys)]
        rows = [[r.get(h, "") for h in headers] for r in obj]
        return headers, rows

    if isinstance(obj, list) and obj and isinstance(obj[0], list):
        first = obj[0]
        if all(isinstance(x, str) for x in first):
            headers = [str(x) for x in first]
            rows = obj[1:]
        else:
            headers = [f"Col{i+1}" for i in range(len(first))]
            rows = obj
        return headers, rows

    return [], []


def _style_header_row(ws, row_idx: int = 1) -> None:
    for cell in ws[row_idx]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def _autosize_columns(ws) -> None:
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[letter]:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_len:
                max_len = len(value)
        ws.column_dimensions[letter].width = min(max(12, max_len + 2), 60)


def _apply_fixed_widths(ws, fixed_widths: Dict[str, float]) -> None:
    header_map = {str(ws.cell(row=1, column=idx).value or ""): get_column_letter(idx) for idx in range(1, ws.max_column + 1)}
    for header, width in fixed_widths.items():
        letter = header_map.get(header)
        if letter:
            ws.column_dimensions[letter].width = width


def _estimate_line_count(text: str, width: float) -> int:
    if not text:
        return 1
    usable_width = max(8, int(width) - 2)
    total = 0
    for line in text.splitlines() or [text]:
        total += max(1, math.ceil(max(1, len(line)) / usable_width))
    return total


def _apply_row_heights(ws, min_height: float = 20.0, line_height: float = 15.0) -> None:
    for row_idx in range(1, ws.max_row + 1):
        max_lines = 1
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            text = "" if cell.value is None else str(cell.value)
            width = float(ws.column_dimensions[get_column_letter(col_idx)].width or 12)
            max_lines = max(max_lines, _estimate_line_count(text, width))
        ws.row_dimensions[row_idx].height = max(min_height, min(240.0, max_lines * line_height))


def _style_data_region(ws, recommended_cols: List[str], editable_cols: List[str]) -> None:
    header_map = {str(ws.cell(row=1, column=idx).value or ""): idx for idx in range(1, ws.max_column + 1)}
    for col_name in recommended_cols:
        idx = header_map.get(col_name)
        if not idx:
            continue
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=idx).fill = RECOMMENDED_FILL
    for col_name in editable_cols:
        idx = header_map.get(col_name)
        if not idx:
            continue
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=idx).fill = EDITABLE_FILL
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def _write_table_sheet(
    ws,
    headers: List[str],
    rows: List[List[Any]],
    freeze: str = "A2",
    fixed_widths: Dict[str, float] | None = None,
) -> None:
    ws.append(headers)
    _style_header_row(ws)
    for row in rows:
        ws.append(row)
    ws.freeze_panes = freeze
    if headers:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(1, ws.max_row)}"
    _autosize_columns(ws)
    if fixed_widths:
        _apply_fixed_widths(ws, fixed_widths)
    _apply_row_heights(ws)


def _build_readme_rows(contract: Dict[str, Any]) -> List[List[str]]:
    metadata_rows = [
        ["run_id", str(contract.get("run_id") or "")],
        ["generated_at", str(contract.get("generated_at") or "")],
        ["source_endpoint", str(contract.get("source_endpoint") or "")],
    ]
    return metadata_rows + [
        ["section", "details"],
        ["Workbook purpose", "Use this workbook to review the grain worker's structural recommendations before later specialists continue. The goal is to confirm the base row grain, validate repeat families, and capture any naming or structural overrides."],
        ["Primary Grain", "The minimal key or key combination that should identify one row in the base table."],
        ["Candidate Dimension", "A stable grouping or entity that may deserve its own dimension-style table, but is not the primary row grain."],
        ["Repeat Family", "A group of repeated or matrix-style columns that may need to become a child table instead of staying wide."],
        ["How to edit", "Do not overwrite recommended columns. Only edit status, your_* columns, and comments on the editable sheets."],
        ["Status: accept", "Keep the recommendation exactly as shown. You do not need to fill the your_* columns."],
        ["Status: modify", "Replace the recommended grain with your own key columns by filling the your_* columns. Leave the recommended columns untouched."],
        ["Status: unsure", "Do not finalize the recommendation yet. Leave comments explaining the uncertainty or what extra review is needed."],
        ["Modify scope", "In this light contract, modify is mainly intended for the Primary Grain sheet. For families and dimensions, prefer accept, reject, unsure, and comments."],
        ["Composite key guidance", "To create a composite key, fill your_key_1 and then add any additional key parts in your_key_2 and your_key_3. Leave unused trailing key cells blank. Example: a two-part key uses only your_key_1 and your_key_2."],
        ["3-step workflow", "1) Review the recommendation sheets. 2) Set status for each row. 3) Fill the your_* fields only when modifying."],
    ]


def build_xlsx_bytes(rows_table_json: str, sheet_name: str = "Template") -> bytes:
    parsed = json.loads(rows_table_json)
    headers, rows = _normalize_rows(parsed)

    wb = Workbook()
    ws = wb.active
    ws.title = (sheet_name[:31] if sheet_name else "Template")
    if headers:
        _write_table_sheet(ws, headers, rows)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_light_contract_xlsx_bytes(contract: Dict[str, Any]) -> bytes:
    wb = Workbook()

    # Sheet 1: Read Me
    ws = wb.active
    ws.title = "Read Me"
    readme_rows = _build_readme_rows(contract)
    ws.append(["section", "details"])
    _style_header_row(ws)
    for row in readme_rows:
        ws.append(row)
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 72
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    _apply_row_heights(ws)

    # Sheet 2: Column Guide
    ws_cols = wb.create_sheet(title="Column Guide")
    column_headers = ["column_index", "column_name", "family_id", "notes"]
    column_rows = [
        [
            row.get("column_index", ""),
            row.get("column_name", ""),
            row.get("family_id", ""),
            row.get("notes", ""),
        ]
        for row in contract.get("column_guide_rows", [])
    ]
    _write_table_sheet(ws_cols, column_headers, column_rows, fixed_widths={"column_name": 28, "family_id": 18, "notes": 40})

    # Sheet 3: Grain Summary
    ws_summary = wb.create_sheet(title="Grain Summary")
    summary_headers = ["topic", "recommendation", "why", "needs_review"]
    summary_rows = [
        [
            row.get("topic", ""),
            row.get("recommendation", ""),
            row.get("why", ""),
            row.get("needs_review", ""),
        ]
        for row in contract.get("grain_summary_rows", [])
    ]
    _write_table_sheet(ws_summary, summary_headers, summary_rows, fixed_widths={"topic": 22, "recommendation": 30, "why": 56, "needs_review": 14})

    # Sheet 4: Primary Grain
    ws_grain = wb.create_sheet(title="Primary Grain")
    grain_headers = [
        "item",
        "recommended_key_1",
        "recommended_key_2",
        "recommended_key_3",
        "your_key_1",
        "your_key_2",
        "your_key_3",
        "status",
        "comments",
    ]
    grain_rows = [
        [
            row.get("item", ""),
            row.get("recommended_key_1", ""),
            row.get("recommended_key_2", ""),
            row.get("recommended_key_3", ""),
            row.get("your_key_1", ""),
            row.get("your_key_2", ""),
            row.get("your_key_3", ""),
            row.get("status", ""),
            row.get("comments", ""),
        ]
        for row in contract.get("primary_grain_rows", [])
    ]
    _write_table_sheet(ws_grain, grain_headers, grain_rows, fixed_widths={"item": 18, "comments": 50})
    _style_data_region(
        ws_grain,
        recommended_cols=["recommended_key_1", "recommended_key_2", "recommended_key_3"],
        editable_cols=["your_key_1", "your_key_2", "your_key_3", "status", "comments"],
    )

    # Sheet 5: Dimension Tables
    ws_dim = wb.create_sheet(title="Dimension Tables")
    dim_headers = [
        "table_name",
        "recommended_key_1",
        "recommended_key_2",
        "recommended_key_3",
        "your_key_1",
        "your_key_2",
        "your_key_3",
        "relationship_to_primary",
        "status",
        "comments",
    ]
    dim_rows = [
        [
            row.get("table_name", ""),
            row.get("recommended_key_1", ""),
            row.get("recommended_key_2", ""),
            row.get("recommended_key_3", ""),
            row.get("your_key_1", ""),
            row.get("your_key_2", ""),
            row.get("your_key_3", ""),
            row.get("relationship_to_primary", ""),
            row.get("status", ""),
            row.get("comments", ""),
        ]
        for row in contract.get("dimension_rows", [])
    ]
    _write_table_sheet(ws_dim, dim_headers, dim_rows, fixed_widths={"table_name": 28, "relationship_to_primary": 18, "comments": 48})
    _style_data_region(
        ws_dim,
        recommended_cols=["recommended_key_1", "recommended_key_2", "recommended_key_3", "relationship_to_primary"],
        editable_cols=["your_key_1", "your_key_2", "your_key_3", "status", "comments"],
    )

    # Sheet 6: Repeat Families
    ws_fam = wb.create_sheet(title="Repeat Families")
    fam_headers = [
        "family_id",
        "recommended_table_name",
        "your_table_name",
        "recommended_repeat_index_name",
        "your_repeat_index_name",
        "recommended_parent_key",
        "your_parent_key",
        "status",
        "comments",
    ]
    fam_rows = [
        [
            row.get("family_id", ""),
            row.get("recommended_table_name", ""),
            row.get("your_table_name", ""),
            row.get("recommended_repeat_index_name", ""),
            row.get("your_repeat_index_name", ""),
            row.get("recommended_parent_key", ""),
            row.get("your_parent_key", ""),
            row.get("status", ""),
            row.get("comments", ""),
        ]
        for row in contract.get("repeat_family_rows", [])
    ]
    _write_table_sheet(ws_fam, fam_headers, fam_rows, fixed_widths={"family_id": 18, "recommended_table_name": 28, "recommended_repeat_index_name": 20, "recommended_parent_key": 24, "comments": 56})
    _style_data_region(
        ws_fam,
        recommended_cols=["recommended_table_name", "recommended_repeat_index_name", "recommended_parent_key"],
        editable_cols=["your_table_name", "your_repeat_index_name", "your_parent_key", "status", "comments"],
    )

    structural_gate_rows = contract.get("structural_gate_rows", [])
    if structural_gate_rows:
        ws_gates = wb.create_sheet(title="Structural Gates")
        gate_headers = [
            "trigger_column",
            "trigger_value",
            "affected_column_count",
            "affected_family_ids",
            "missing_explained_pct",
            "directionality",
            "interpretation",
        ]
        gate_rows = [
            [
                row.get("trigger_column", ""),
                row.get("trigger_value", ""),
                row.get("affected_column_count", ""),
                ", ".join(row.get("affected_family_ids", []) or []),
                row.get("missing_explained_pct", ""),
                row.get("directionality", ""),
                row.get("interpretation", ""),
            ]
            for row in structural_gate_rows
        ]
        _write_table_sheet(ws_gates, gate_headers, gate_rows, fixed_widths={"trigger_column": 24, "trigger_value": 18, "affected_family_ids": 32, "interpretation": 54})

    # Sheet 7/8: Overrides
    ws_over = wb.create_sheet(title="Overrides")
    override_headers = ["field", "description", "user_input"]
    override_rows = [
        [
            row.get("field", ""),
            row.get("description", ""),
            row.get("user_input", ""),
        ]
        for row in contract.get("override_rows", [])
    ]
    _write_table_sheet(ws_over, override_headers, override_rows, fixed_widths={"field": 34, "description": 58, "user_input": 58})
    _style_data_region(ws_over, recommended_cols=["field", "description"], editable_cols=["user_input"])

    grain_status = DataValidation(type="list", formula1='"accept,modify,unsure"', allow_blank=True)
    ws_grain.add_data_validation(grain_status)
    grain_status.add(f"H2:H{max(2, ws_grain.max_row)}")

    dim_status = DataValidation(type="list", formula1='"accept,reject,unsure"', allow_blank=True)
    ws_dim.add_data_validation(dim_status)
    dim_status.add(f"I2:I{max(2, ws_dim.max_row)}")

    fam_status = DataValidation(
        type="list",
        formula1='"accept,reject,unsure"',
        allow_blank=True,
    )
    ws_fam.add_data_validation(fam_status)
    fam_status.add(f"H2:H{max(2, ws_fam.max_row)}")

    for sheet in wb.worksheets:
        _apply_row_heights(sheet)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _sheet_rows_as_dicts(ws) -> List[Dict[str, Any]]:
    if ws.max_row < 2 or ws.max_column < 1:
        return []
    headers = [str(ws.cell(row=1, column=idx).value or "").strip() for idx in range(1, ws.max_column + 1)]
    rows: List[Dict[str, Any]] = []
    for row_idx in range(2, ws.max_row + 1):
        values = [ws.cell(row=row_idx, column=idx).value for idx in range(1, ws.max_column + 1)]
        if not any(v not in (None, "") for v in values):
            continue
        rows.append({headers[idx - 1]: values[idx - 1] for idx in range(1, ws.max_column + 1) if headers[idx - 1]})
    return rows


def parse_light_contract_xlsx_bytes(xlsx_bytes: bytes) -> Dict[str, Any]:
    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)

    def _sheet(name: str):
        if name not in wb.sheetnames:
            return None
        return wb[name]

    readme_ws = _sheet("Read Me")
    readme_rows = _sheet_rows_as_dicts(readme_ws) if readme_ws else []
    metadata: Dict[str, str] = {}
    for row in readme_rows:
        section = str(row.get("section") or "").strip()
        details = "" if row.get("details") is None else str(row.get("details"))
        if section in {"run_id", "generated_at", "source_endpoint"}:
            metadata[section] = details

    primary_rows = _sheet_rows_as_dicts(_sheet("Primary Grain")) if _sheet("Primary Grain") else []
    dimension_rows = _sheet_rows_as_dicts(_sheet("Dimension Tables")) if _sheet("Dimension Tables") else []
    family_rows = _sheet_rows_as_dicts(_sheet("Repeat Families")) if _sheet("Repeat Families") else []
    override_rows = _sheet_rows_as_dicts(_sheet("Overrides")) if _sheet("Overrides") else []

    return {
        "metadata": metadata,
        "primary_grain_rows": primary_rows,
        "dimension_rows": dimension_rows,
        "repeat_family_rows": family_rows,
        "override_rows": override_rows,
    }
